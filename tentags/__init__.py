"""
TenTags
=======

Declarative Document DSL.

Write a document once.
Render it everywhere.

HTML • PDF • XLSX

Quick Start
-----------

    import tentags
    html = tentags.render('3,3,1,"black","solid",0, data(A,B,C; D,E,F; G,H,I)')

Typical Workflow
----------------

1. Define style.
2. Define data.
3. Compile into TableModel.
4. Render to HTML, PDF or XLSX.

or simply:

    html = tentags.render(formula)

API Index
---------

Main API:
  render()           Render HTML
  compile()          Build TableModel
  parse()            Parse formula
  validate()         Validate syntax

Serializer API:
  serialize          Namespace for preamble/style/data serializers
  dumps_preamble()   Serialize Python values to a TenTags preamble
  dumps_style()      Serialize a Python matrix to style(...)
  dumps_data()       Serialize a Python matrix to data(...)

Export:
  render_html()      Render to HTML string
  render_pdf()       Export to PDF document
  render_xlsx()      Export to XLSX sheet

Multiple Tables:
  multitable_html()  Combine tables into HTML Grid
  multitable_pdf()   Combine tables into multi-page PDF
  multitable_xlsx()  Combine tables into Excel workbook

Utilities:
  demo()             Generate sample documents
  get_prompt()       Return the bundled LLM bootstrap prompt
  info()             Display system diagnostic info
  features()         Check available render backends

Website: https://tentags.org
Documentation: https://tentags.org/docs
GitHub: https://github.com/Jandos77/tentags

Current Version: 2.1.8
License: Apache License 2.0
"""

__version__ = "2.1.8"
__author__ = "Zhandos Mambetali"
__license__ = "Apache-2.0"
__copyright__ = "Copyright (c) 2026 Zhandos Mambetali"
__homepage__ = "https://tentags.org"
__url__ = "https://tentags.org"
version_info = (2, 1, 8)

__all__ = [
    "__version__",
    "__author__",
    "__license__",
    "__copyright__",
    "__homepage__",
    "__url__",
    "version_info",
    "parse",
    "render",
    "render_html",
    "render_xlsx",
    "render_pdf",
    "compile",
    "serialize",
    "dumps_preamble",
    "dumps_style",
    "dumps_data",
    "multitable_html",
    "multitable_xlsx",
    "multitable_pdf",
    "DEFAULT_MULTITABLE_HTML_SETTINGS",
    "DEFAULT_MULTITABLE_XLSX_SETTINGS",
    "DEFAULT_MULTITABLE_PDF_SETTINGS",
    "features",
    "info",
    "get_prompt",
    "validate",
    "demo",
    "build_mark_index",
    "AddressContext",
    "AddressResolver",
    "AddressTarget",
    "ResolvedAddress",
    "TableModel",
    "CellDesc",
    "BorderFlags",
    "ScaleError"
]

import re as _re
import csv as _csv
import urllib.request as _urllib_request
import io as _io
import os as _os
import html as _html
from typing import Union as _Union, Any as _Any, Optional as _Optional, Dict as _Dict, List as _List
from enum import Enum as _Enum, auto as _auto
from dataclasses import dataclass as _dataclass
from tentags.addressing import AddressType as _AddressType
from tentags.addressing import AddressContext
from tentags.addressing import AddressResolver
from tentags.addressing import AddressTarget
from tentags.addressing import column_to_name as _column_to_name
from tentags.addressing import parse_cell_ref as _parse_cell_ref
from tentags.addressing import DuplicateMarkError as _DuplicateMarkError
from tentags.addressing import parse_address as _parse_address
from tentags.addressing import ResolvedAddress

_SENTINEL = object()

class ScaleError(ValueError):
    """Raised when a preamble scale(...) extension is invalid."""


DEFAULT_MULTITABLE_HTML_SETTINGS = {
    "output": None,
    "table_order": None,
    "columns": None,
    "tables_per_row": None,
    "html_title": "Multi-Table Report",
    "layout": "vertical",
    "cols": 1,
    "gap": "24px",
    "full_page": False,
}

DEFAULT_MULTITABLE_XLSX_SETTINGS = {
    "output": None,
    "table_order": None,
    "columns": None,
    "tables_per_sheet": None,
    "stacked_sheet_name": "Report",
    "mode": "sheets",
    "gap": 3,
    "show_titles": True,
}

DEFAULT_MULTITABLE_PDF_SETTINGS = {
    "output": None,
    "table_order": None,
    "columns": None,
    "tables_per_row": 1,
    "tables_per_page": None,
    "gap": 12,
    "page_size": "letter",
    "orientation": "portrait",
    "page_break_after_each": True,
    "margins": (36, 36, 36, 36),
}

def _dumps_quoted(value: _Any) -> str:
    text = str(value)
    text = text.replace("\\", "\\\\").replace('"', '\\"')
    return f'"{text}"'

def _require_non_negative_int(value: _Any, name: str) -> int:
    if isinstance(value, bool) or not isinstance(value, int) or value < 0:
        raise ValueError(f"{name} must be a non-negative integer.")
    return value

def _require_positive_int(value: _Any, name: str) -> int:
    if isinstance(value, bool) or not isinstance(value, int) or value < 1:
        raise ValueError(f"{name} must be a positive integer.")
    return value

def _parse_scale_block(scale_block: str, rows: int, cols: int) -> tuple[dict[int, int], dict[int, int]]:
    match = _re.fullmatch(r"(?is)scale\((.*)\)", str(scale_block).strip())
    if not match:
        raise ScaleError("Invalid scale syntax. Expected scale(A1=1,3;C5=2,2).")

    body = match.group(1).strip()
    if not body:
        raise ScaleError("scale(...) requires at least one A1=vertical,horizontal entry.")

    entries = [entry.strip() for entry in body.split(";")]
    if entries and entries[-1] == "":
        entries.pop()
    if not entries or any(not entry for entry in entries):
        raise ScaleError("scale(...) contains an empty entry.")

    row_scales = {}
    col_scales = {}
    entry_pattern = _re.compile(r"^([A-Za-z]+[1-9][0-9]*)\s*=\s*([1-5])\s*,\s*([1-5])$")

    for entry in entries:
        entry_match = entry_pattern.fullmatch(entry)
        if not entry_match:
            raise ScaleError(
                f"Invalid scale entry {entry!r}. Expected A1=v,h with integer values from 1 to 5."
            )

        address_text, vertical_text, horizontal_text = entry_match.groups()
        try:
            cell = _parse_cell_ref(address_text)
        except ValueError as exc:
            raise ScaleError(f"Invalid scale cell address {address_text!r}.") from exc

        if cell.row >= rows or cell.col >= cols:
            raise ScaleError(
                f"Scale address {address_text.upper()} is outside the {rows}x{cols} table."
            )

        vertical = int(vertical_text)
        horizontal = int(horizontal_text)
        if vertical > 1:
            row_scales[cell.row] = max(row_scales.get(cell.row, 1), vertical)
        if horizontal > 1:
            col_scales[cell.col] = max(col_scales.get(cell.col, 1), horizontal)

    return row_scales, col_scales

def _serialize_scale_mapping(scale: _Any, rows: int, cols: int) -> str:
    if scale is None:
        return ""
    if not isinstance(scale, dict):
        raise ScaleError("scale must be a dict such as {'A1': (1, 3)}.")
    if not scale:
        return ""

    normalized = {}
    for address_text, values in scale.items():
        try:
            cell = _parse_cell_ref(str(address_text))
        except ValueError as exc:
            raise ScaleError(f"Invalid scale cell address {address_text!r}.") from exc
        if cell.row >= rows or cell.col >= cols:
            canonical = f"{_column_to_name(cell.col)}{cell.row + 1}"
            raise ScaleError(f"Scale address {canonical} is outside the {rows}x{cols} table.")
        if not isinstance(values, (tuple, list)) or len(values) != 2:
            raise ScaleError(f"Scale value for {address_text!r} must be a pair (vertical, horizontal).")
        vertical, horizontal = values
        if any(isinstance(value, bool) or not isinstance(value, int) or value < 1 or value > 5 for value in values):
            raise ScaleError(f"Scale value for {address_text!r} must contain integers from 1 to 5.")

        key = (cell.row, cell.col)
        previous = normalized.get(key, (1, 1))
        normalized[key] = (max(previous[0], vertical), max(previous[1], horizontal))

    entries = [
        f"{_column_to_name(col)}{row + 1}={vertical},{horizontal}"
        for (row, col), (vertical, horizontal) in sorted(normalized.items())
    ]
    return f"scale({';'.join(entries)})"

def dumps_preamble(
    rows: int,
    cols: int,
    border_width: int = 1,
    border_color: str = "#cbd5e1",
    border_style: str = "solid",
    stretch: int = 0,
    cell_height: int = 30,
    scale: _Any = None,
) -> str:
    """
    Serializes Python preamble values into canonical TenTags preamble text.

    This is a serializer only. It does not compile or create IR.
    """
    rows = _require_positive_int(rows, "rows")
    cols = _require_positive_int(cols, "cols")
    border_width = _require_non_negative_int(border_width, "border_width")
    stretch = _require_non_negative_int(stretch, "stretch")
    cell_height = _require_non_negative_int(cell_height, "cell_height")
    parts = [
        str(rows),
        str(cols),
        str(border_width),
        _dumps_quoted(border_color),
        _dumps_quoted(border_style),
        str(stretch),
        str(cell_height),
    ]
    scale_block = _serialize_scale_mapping(scale, rows, cols)
    if scale_block:
        parts.append(scale_block)
    return ",".join(parts)

def _coerce_serializer_matrix(rows: _Any, name: str) -> list[list[_Any]]:
    if rows is None or isinstance(rows, (str, bytes)):
        raise ValueError(f"{name} must be a matrix: a list of rows.")
    try:
        row_list = list(rows)
    except TypeError as exc:
        raise ValueError(f"{name} must be a matrix: a list of rows.") from exc

    matrix = []
    for index, row in enumerate(row_list):
        if row is None or isinstance(row, (str, bytes)):
            raise ValueError(f"{name} row {index + 1} must be a list of cells.")
        try:
            matrix.append(list(row))
        except TypeError as exc:
            raise ValueError(f"{name} row {index + 1} must be a list of cells.") from exc
    return matrix

def _validate_serializer_matrix(
    matrix: list[list[_Any]],
    name: str,
    expected_rows: int = None,
    expected_cols: int = None,
) -> None:
    if expected_rows is not None:
        _require_non_negative_int(expected_rows, "expected_rows")
        if len(matrix) != expected_rows:
            raise ValueError(f"{name} expected {expected_rows} rows, got {len(matrix)}.")
    if expected_cols is not None:
        _require_non_negative_int(expected_cols, "expected_cols")
        for index, row in enumerate(matrix):
            if len(row) != expected_cols:
                raise ValueError(
                    f"{name} row {index + 1} expected {expected_cols} cells, got {len(row)}."
                )

def _dumps_cell(value: _Any) -> str:
    if value is None:
        return ""
    return str(value)

def _dumps_block(block_name: str, rows: _Any, expected_rows: int = None, expected_cols: int = None) -> str:
    matrix = _coerce_serializer_matrix(rows, block_name)
    _validate_serializer_matrix(matrix, block_name, expected_rows, expected_cols)
    body = ";\n".join(", ".join(_dumps_cell(cell) for cell in row) for row in matrix)
    return f"{block_name}(\n{body}\n)"

def dumps_style(rows: _Any, expected_rows: int = None, expected_cols: int = None) -> str:
    """
    Serializes a Python matrix into a TenTags style(...) block.

    Cell values are raw TenTags style expressions. None becomes an empty cell.
    """
    return _dumps_block("style", rows, expected_rows, expected_cols)

def dumps_data(rows: _Any, expected_rows: int = None, expected_cols: int = None) -> str:
    """
    Serializes a Python matrix into a TenTags data(...) block.

    Cell values are raw TenTags data expressions. None becomes an empty cell.
    """
    return _dumps_block("data", rows, expected_rows, expected_cols)

class _SerializeNamespace:
    """
    Namespace for serializers that convert Python structures to TenTags DSL.
    """
    @staticmethod
    def preamble(
        rows: int,
        cols: int,
        border_width: int = 1,
        border_color: str = "#cbd5e1",
        border_style: str = "solid",
        stretch: int = 0,
        cell_height: int = 30,
        scale: _Any = None,
    ) -> str:
        return dumps_preamble(
            rows=rows,
            cols=cols,
            border_width=border_width,
            border_color=border_color,
            border_style=border_style,
            stretch=stretch,
            cell_height=cell_height,
            scale=scale,
        )

    @staticmethod
    def style(rows: _Any, expected_rows: int = None, expected_cols: int = None) -> str:
        return dumps_style(rows, expected_rows=expected_rows, expected_cols=expected_cols)

    @staticmethod
    def data(rows: _Any, expected_rows: int = None, expected_cols: int = None) -> str:
        return dumps_data(rows, expected_rows=expected_rows, expected_cols=expected_cols)

serialize = _SerializeNamespace()

class _TokenType(_Enum):
    TAG_OPEN = _auto()
    TAG_CLOSE = _auto()
    TAG_SELF = _auto()
    COMMA = _auto()
    SEMICOLON = _auto()
    TEXT = _auto()

@_dataclass
class _Token:
    type: _TokenType
    value: str  # tag name or text value
    line: int
    column: int
    attr: str = None  # optional attribute value (e.g. tag parameter in <tag=value>)

@_dataclass
class _Link:
    scheme: str
    target: _Any
    raw: str

class BorderFlags:
    NONE = 0
    HIDE_LEFT = 1
    HIDE_RIGHT = 2
    HIDE_TOP = 4
    HIDE_BOTTOM = 8

class CellDesc:
    def __init__(self):
        self.raw_expr = ""
        self.cm_block_id = None
        self.rm_block_id = None
        self.border_flags = BorderFlags.NONE
        self.text_parts = []
        self.images = []
        self.link = None
        self.mark = None
        self.value_refs = []
        self.styles = {}  # Extensible styles, e.g. {'font-weight': 'bold', 'color': '#ff0000'}

def _is_self_tag(tag_name: str) -> bool:
    return tag_name.lower() in ('img', 'mark', 'value')

def _parse_link(attr: str) -> _Link:
    raw = str(attr or "").strip()
    if raw.lower().startswith("goto:"):
        target = _parse_address(raw[5:].strip())
        return _Link(scheme="goto", target=target, raw=raw)

    scheme_match = _re.match(r"^([A-Za-z][A-Za-z0-9+.-]*):", raw)
    scheme = scheme_match.group(1).lower() if scheme_match else "url"
    return _Link(scheme=scheme, target=raw, raw=raw)

def _html_cell_id(row: int, col: int, prefix: str = None) -> str:
    cell_name = f"{_column_to_name(col)}{row + 1}"
    return f"tt-{prefix}-{cell_name}" if prefix else f"tt-{cell_name}"

def _html_mark_id(mark: str, prefix: str = None) -> str:
    safe = _re.sub(r"[^A-Za-z0-9_.:-]+", "-", str(mark).strip())
    safe = safe.strip("-") or "mark"
    return f"tt-{prefix}-mark-{safe}" if prefix else f"tt-mark-{safe}"

def _make_address_target(
    model: "TableModel",
    *,
    project: str = None,
    document: str = None,
    table: str = None,
    sheet: str = None,
    list_name: str = None,
    html_prefix: str = None,
    xlsx_sheet_name: str = "Table",
    xlsx_start_row: int = 1,
    pdf_prefix: str = None,
) -> AddressTarget:
    return AddressTarget(
        model=model,
        context=AddressContext(project=project, document=document, table=table, sheet=sheet, list_name=list_name),
        html_prefix=html_prefix,
        xlsx_sheet_name=xlsx_sheet_name,
        xlsx_start_row=xlsx_start_row,
        pdf_prefix=pdf_prefix,
    )

def _local_address_resolver(current: AddressTarget) -> AddressResolver:
    return AddressResolver([current])

def _address_to_html_href(
    address,
    resolver: AddressResolver = None,
    current: AddressTarget = None,
    prefix_attr: str = "html_prefix",
) -> str:
    resolved = resolver.resolve(address, current) if resolver is not None else None
    if resolved is not None:
        cell = resolved.start_cell()
        prefix = getattr(resolved.target, prefix_attr)
        if address.location.type is _AddressType.MARK:
            return f"#{_html_mark_id(address.location.mark, prefix)}"
        if cell is not None:
            return f"#{_html_cell_id(cell.row, cell.col, prefix)}"
        return "#"

    location = address.location
    if location.type is _AddressType.CELL:
        return f"#{_html_cell_id(location.cell.row, location.cell.col)}"
    if location.type is _AddressType.RANGE:
        return f"#{_html_cell_id(location.range.start.row, location.range.start.col)}"
    if location.type is _AddressType.MARK:
        return f"#{_html_mark_id(location.mark)}"
    return "#"

def _link_to_html_href(
    link: _Link,
    resolver: AddressResolver = None,
    current: AddressTarget = None,
    prefix_attr: str = "html_prefix",
) -> str:
    if link.scheme == "goto":
        return _address_to_html_href(link.target, resolver, current, prefix_attr)
    return str(link.target)

def _xlsx_sheet_ref(sheet_name: str) -> str:
    if _re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", sheet_name):
        return sheet_name
    return "'" + sheet_name.replace("'", "''") + "'"

def _address_to_xlsx_hyperlink(
    address,
    sheet_name: str,
    mark_index: dict = None,
    resolver: AddressResolver = None,
    current: AddressTarget = None,
) -> str:
    resolved = resolver.resolve(address, current) if resolver is not None else None
    if resolved is not None:
        cell = resolved.start_cell()
        if cell is None:
            return None
        target_sheet = resolved.target.xlsx_sheet_name or sheet_name
        row = cell.row + max(1, resolved.target.xlsx_start_row)
        cell_name = f"{_column_to_name(cell.col)}{row}"
        return f"#{_xlsx_sheet_ref(target_sheet)}!{cell_name}"

    location = address.location
    if location.type is _AddressType.CELL:
        cell_name = f"{_column_to_name(location.cell.col)}{location.cell.row + 1}"
        return f"#{_xlsx_sheet_ref(sheet_name)}!{cell_name}"
    if location.type is _AddressType.RANGE:
        cell = location.range.start
        cell_name = f"{_column_to_name(cell.col)}{cell.row + 1}"
        return f"#{_xlsx_sheet_ref(sheet_name)}!{cell_name}"
    if location.type is _AddressType.MARK and mark_index:
        coords = mark_index.get(location.mark)
        if coords is not None:
            return f"#{_xlsx_sheet_ref(sheet_name)}!{_column_to_name(coords[1])}{coords[0] + 1}"
    return None

def _link_to_xlsx_hyperlink(
    link: _Link,
    sheet_name: str,
    mark_index: dict = None,
    resolver: AddressResolver = None,
    current: AddressTarget = None,
) -> str:
    if link.scheme == "goto":
        return _address_to_xlsx_hyperlink(link.target, sheet_name, mark_index, resolver, current)
    return str(link.target)

def build_mark_index(model: "TableModel") -> dict:
    marks = {}
    for row_index, row in enumerate(model.cells):
        for col_index, cell in enumerate(row):
            if not getattr(cell, "mark", None):
                continue
            mark = str(cell.mark)
            if mark in marks:
                raise _DuplicateMarkError(f"Duplicate mark: {mark}")
            marks[mark] = (row_index, col_index)
    return marks

def _has_external_scope(address) -> bool:
    return bool(address.project or address.document or address.table)

def _cell_value_at(model: "TableModel", row: int, col: int) -> str:
    if row < 0 or col < 0:
        return ""
    if row >= len(model.cells) or col >= len(model.cells[row]):
        return ""
    value = model.cells[row][col].raw_expr
    return "" if value == "None" else str(value)

def _resolve_value_address(model: "TableModel", address, mark_index: dict = None) -> str:
    if _has_external_scope(address):
        raise ValueError("External <value=...> addresses are not supported yet.")

    location = address.location
    if location.type is _AddressType.CELL:
        return _cell_value_at(model, location.cell.row, location.cell.col)

    if location.type is _AddressType.RANGE:
        values = [
            _cell_value_at(model, cell.row, cell.col)
            for cell in location.range.iter_cells()
        ]
        return ", ".join(values)

    if location.type is _AddressType.MARK:
        if mark_index is None:
            mark_index = build_mark_index(model)
        coords = mark_index.get(location.mark)
        if coords is None:
            return ""
        return _cell_value_at(model, coords[0], coords[1])

    return ""

def _resolve_value_refs(model: "TableModel") -> "TableModel":
    refs = [
        (cell, placeholder, address)
        for row in model.cells
        for cell in row
        for placeholder, address in getattr(cell, "value_refs", [])
    ]
    if not refs:
        return model

    needs_mark_index = any(
        address.location.type is _AddressType.MARK
        for _, _, address in refs
    )
    mark_index = build_mark_index(model) if needs_mark_index else None

    for cell, placeholder, address in refs:
        value = _resolve_value_address(model, address, mark_index)
        cell.raw_expr = cell.raw_expr.replace(placeholder, value)
        cell.text_parts = [
            part.replace(placeholder, value) if isinstance(part, str) else part
            for part in cell.text_parts
        ]
    return model

class TableModel:
    def __init__(
        self,
        rows: int,
        cols: int,
        cells: list[list[CellDesc]],
        border_width: int,
        border_color: str,
        border_style: str,
        stretch: int,
        cell_height: int,
        row_scales: dict[int, int] = None,
        col_scales: dict[int, int] = None,
    ):
        self.rows = rows
        self.cols = cols
        self.cells = cells
        self.border_width = border_width
        self.border_color = border_color
        self.border_style = border_style
        self.stretch = stretch
        self.cell_height = cell_height
        self.row_scales = dict(row_scales or {})
        self.col_scales = dict(col_scales or {})

def _table_item_pycells_table_name(item: _Any) -> str:
    if not isinstance(item, dict):
        return None
    return item.get("document") or item.get("document_name") or item.get("doc") or item.get("workbook")

def _table_item_list_name(item: _Any, index: int) -> str:
    if isinstance(item, dict):
        return (
            item.get("table_name")
            or item.get("list_name")
            or item.get("list")
            or item.get("table")
            or item.get("sheet")
            or item.get("sheet_name")
            or item.get("title")
            or f"Table {index + 1}"
        )
    return f"Table {index + 1}"

def _table_item_model(item: _Any, context: dict = None) -> "TableModel":
    if isinstance(item, TableModel):
        return item
    if isinstance(item, dict):
        if item.get("style") is None and item.get("data") is not None:
            return parse(f'{item.get("preamble")}, {item.get("data")}', context)
        return compile(
            style=item.get("style"),
            data=item.get("data"),
            preamble=item.get("preamble"),
            context=context
        )
    raise TypeError("Each table in 'tables' must be a TableModel or a dict.")

def _table_item_meta(item: _Any, index: int) -> dict:
    table_name = _table_item_pycells_table_name(item)
    list_name = _table_item_list_name(item, index)
    if isinstance(item, dict):
        return {
            "project": item.get("project"),
            "document": table_name,
            "table": list_name,
            "list": list_name,
            "title": item.get("title"),
            "sheet_name": item.get("sheet_name") or list_name,
        }
    return {
        "project": None,
        "document": table_name,
        "table": list_name,
        "list": list_name,
        "title": None,
        "sheet_name": list_name,
    }

def _table_item_key(item: _Any, index: int) -> str:
    meta = _table_item_meta(item, index)
    document = meta.get("document")
    list_name = meta.get("list")
    return f"{document}!{list_name}" if document else str(list_name)

def _merge_settings(defaults: dict, settings: dict = None, overrides: dict = None) -> dict:
    merged = dict(defaults)
    if settings:
        merged.update(settings)
    for key, value in (overrides or {}).items():
        if value is not _SENTINEL:
            merged[key] = value
    return merged

def _prepare_multitable_items(tables: list, settings: dict) -> list:
    ordered = list(tables)
    table_order = settings.get("table_order")
    if not table_order:
        return ordered

    by_key = {}
    for index, item in enumerate(ordered):
        key = _table_item_key(item, index)
        if key in by_key:
            raise ValueError(f"Duplicate multitable key in table_order scope: {key}")
        by_key[key] = item

    missing = [key for key in table_order if key not in by_key]
    if missing:
        raise ValueError(f"Unknown table_order entries: {', '.join(missing)}")
    return [by_key[key] for key in table_order]

def _validate_multitable_columns(materialized: list, settings: dict) -> None:
    columns = settings.get("columns")
    if not columns:
        return
    if not isinstance(columns, dict):
        raise ValueError("multitable columns setting must be a dict keyed by Table!List.")

    for index, (model, meta) in enumerate(materialized):
        key = f"{meta['document']}!{meta['list']}" if meta.get("document") else str(meta["list"])
        expected = columns.get(key)
        if expected is None:
            continue
        actual = [cell.raw_expr for cell in model.cells[0]] if model.cells else []
        if actual != list(expected):
            raise ValueError(
                f"Column settings for {key} do not match the first data row. "
                f"Expected {list(expected)!r}, got {actual!r}."
            )

def _html_settings(defaults: dict, settings: dict = None, overrides: dict = None) -> dict:
    merged = _merge_settings(defaults, settings, overrides)
    tables_per_row = merged.get("tables_per_row")
    if tables_per_row is not None:
        if merged.get("cols") not in (None, defaults.get("cols"), tables_per_row):
            raise ValueError("HTML settings cols and tables_per_row must match.")
        merged["cols"] = tables_per_row
    return merged

def _xlsx_settings(defaults: dict, settings: dict = None, overrides: dict = None) -> dict:
    merged = _merge_settings(defaults, settings, overrides)
    tables_per_sheet = merged.get("tables_per_sheet")
    mode = str(merged.get("mode", "sheets")).lower()
    gap = merged.get("gap")
    if isinstance(gap, str):
        normalized_gap = gap.strip().lower()
        if normalized_gap.endswith(("px", "pt")):
            normalized_gap = normalized_gap[:-2].strip()
        try:
            numeric_gap = float(normalized_gap)
        except ValueError:
            raise ValueError('XLSX gap must be a non-negative integer, or a string like "3", "3px", or "3pt".')
        if not numeric_gap.is_integer():
            raise ValueError('XLSX gap must be a non-negative integer, or a string like "3", "3px", or "3pt".')
        gap = int(numeric_gap)
        merged["gap"] = gap
    elif gap is not None and not isinstance(gap, int):
        raise ValueError('XLSX gap must be a non-negative integer, or a string like "3", "3px", or "3pt".')
    if gap is not None and gap < 0:
        raise ValueError("XLSX gap must be greater than or equal to 0.")
    if tables_per_sheet is not None:
        if mode == "sheets" and tables_per_sheet not in (1, "one", "one_per_sheet"):
            raise ValueError('XLSX mode="sheets" supports tables_per_sheet=1.')
        if mode != "sheets" and tables_per_sheet not in ("all", None):
            raise ValueError('XLSX mode="stacked" currently supports tables_per_sheet="all".')
    merged["mode"] = mode
    return merged

def _pdf_settings(defaults: dict, settings: dict = None, overrides: dict = None) -> dict:
    merged = _merge_settings(defaults, settings, overrides)
    tables_per_row = merged.get("tables_per_row")
    if tables_per_row is not None:
        if isinstance(tables_per_row, str):
            if tables_per_row.lower() != "auto":
                raise ValueError('PDF tables_per_row must be a positive integer or "auto".')
            merged["tables_per_row"] = "auto"
        elif not isinstance(tables_per_row, int) or tables_per_row < 1:
            raise ValueError('PDF tables_per_row must be a positive integer or "auto".')
    tables_per_page = merged.get("tables_per_page")
    if tables_per_page is not None:
        if isinstance(tables_per_page, str):
            if tables_per_page.lower() != "auto":
                raise ValueError('PDF tables_per_page must be a positive integer or "auto".')
            merged["tables_per_page"] = "auto"
        elif not isinstance(tables_per_page, int) or tables_per_page < 1:
            raise ValueError('PDF tables_per_page must be a positive integer or "auto".')
    gap = merged.get("gap")
    if isinstance(gap, str):
        normalized_gap = gap.strip().lower()
        if normalized_gap.endswith(("px", "pt")):
            normalized_gap = normalized_gap[:-2].strip()
        try:
            gap = float(normalized_gap)
        except ValueError:
            raise ValueError('PDF gap must be a number, or a string like "24", "24px", or "24pt".')
        merged["gap"] = gap
    elif gap is not None and not isinstance(gap, (int, float)):
        raise ValueError('PDF gap must be a number, or a string like "24", "24px", or "24pt".')
    if gap is not None and gap < 0:
        raise ValueError("PDF gap must be greater than or equal to 0.")
    if (
        tables_per_page is not None
        and isinstance(merged.get("tables_per_row"), int)
        and isinstance(merged.get("tables_per_page"), int)
        and tables_per_page < merged["tables_per_row"]
    ):
        raise ValueError("PDF tables_per_page must be greater than or equal to tables_per_row.")
    return merged

def _write_html_output(output, html: str) -> None:
    if output is None:
        return
    if hasattr(output, "write"):
        output.write(html)
        return
    with open(output, "w", encoding="utf-8") as file:
        file.write(html)

def _normalize_output_target(output):
    if output is None or hasattr(output, "write"):
        return output
    return _os.fspath(output)

_PDF_FONT_CACHE = None

def _find_existing_file(paths: list[str]) -> str:
    for path in paths:
        if path and _os.path.exists(path):
            return path
    return None

def _pdf_font_names() -> dict:
    """
    Returns ReportLab font names with Unicode coverage when a local TTF is available.
    Falls back to built-in PDF fonts when no suitable TTF can be found.
    """
    global _PDF_FONT_CACHE
    if _PDF_FONT_CACHE is not None:
        return _PDF_FONT_CACHE

    fonts = {
        "regular": "Helvetica",
        "bold": "Helvetica-Bold",
        "italic": "Helvetica-Oblique",
        "bold_italic": "Helvetica-BoldOblique",
    }

    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except Exception:
        _PDF_FONT_CACHE = fonts
        return fonts

    windir = _os.environ.get("WINDIR", r"C:\Windows")
    candidates_regular = [
        _os.path.join(windir, "Fonts", "arial.ttf"),
        _os.path.join(windir, "Fonts", "calibri.ttf"),
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
        "/Library/Fonts/Arial Unicode.ttf",
        "/Library/Fonts/Arial.ttf",
    ]
    candidates_bold = [
        _os.path.join(windir, "Fonts", "arialbd.ttf"),
        _os.path.join(windir, "Fonts", "calibrib.ttf"),
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf",
        "/Library/Fonts/Arial Bold.ttf",
    ]
    candidates_italic = [
        _os.path.join(windir, "Fonts", "ariali.ttf"),
        _os.path.join(windir, "Fonts", "calibrii.ttf"),
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Oblique.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Italic.ttf",
        "/Library/Fonts/Arial Italic.ttf",
    ]
    candidates_bold_italic = [
        _os.path.join(windir, "Fonts", "arialbi.ttf"),
        _os.path.join(windir, "Fonts", "calibriz.ttf"),
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-BoldOblique.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-BoldItalic.ttf",
        "/Library/Fonts/Arial Bold Italic.ttf",
    ]

    regular_path = _find_existing_file(candidates_regular)
    bold_path = _find_existing_file(candidates_bold)
    italic_path = _find_existing_file(candidates_italic)
    bold_italic_path = _find_existing_file(candidates_bold_italic)
    if regular_path:
        try:
            pdfmetrics.registerFont(TTFont("TenTagsUnicode", regular_path))
            fonts["regular"] = "TenTagsUnicode"
            fonts["italic"] = "TenTagsUnicode"
            fonts["bold_italic"] = "TenTagsUnicode"
        except Exception:
            pass
    if bold_path and fonts["regular"] == "TenTagsUnicode":
        try:
            pdfmetrics.registerFont(TTFont("TenTagsUnicode-Bold", bold_path))
            fonts["bold"] = "TenTagsUnicode-Bold"
        except Exception:
            fonts["bold"] = fonts["regular"]
    if italic_path and fonts["regular"] == "TenTagsUnicode":
        try:
            pdfmetrics.registerFont(TTFont("TenTagsUnicode-Italic", italic_path))
            fonts["italic"] = "TenTagsUnicode-Italic"
        except Exception:
            fonts["italic"] = fonts["regular"]
    if bold_italic_path and fonts["regular"] == "TenTagsUnicode":
        try:
            pdfmetrics.registerFont(TTFont("TenTagsUnicode-BoldItalic", bold_italic_path))
            fonts["bold_italic"] = "TenTagsUnicode-BoldItalic"
        except Exception:
            fonts["bold_italic"] = fonts["bold"]

    _PDF_FONT_CACHE = fonts
    return fonts

def _scan_open_tag(source: str):
    tag_open_match = _re.match(r'^<([a-zA-Z_]+)(?:=([^>]+)|\s+([^>]+))?>', source)
    if not tag_open_match:
        return None
    tag_name = tag_open_match.group(1)
    tag_attr = tag_open_match.group(2) if tag_open_match.group(2) is not None else tag_open_match.group(3)
    return tag_open_match.group(0), tag_name, tag_attr

def _parse_tag_attrs(attr_str: str, line: int, column: int) -> dict:
    attrs = {}
    if not attr_str:
        return attrs
    for part in attr_str.split():
        if '=' not in part:
            raise ValueError(f"Invalid tag attribute '{part}' at line {line}, column {column}. Expected key=value.")
        key, value = part.split('=', 1)
        key = key.strip().lower()
        value = value.strip().strip('"\'')
        if not key or not value:
            raise ValueError(f"Invalid tag attribute '{part}' at line {line}, column {column}. Expected key=value.")
        attrs[key] = value
    return attrs

def _normalize_img_dimension(value: str, attr_name: str, line: int, column: int, allow_auto: bool = True) -> str:
    value = value.strip().strip('"\'')
    if allow_auto and value.lower() == 'auto':
        return 'auto'
    if _re.fullmatch(r'\d+', value):
        return value
    allowed = "a pixel number or auto" if allow_auto else "a pixel number"
    raise ValueError(f"<img> attribute {attr_name}= must be {allowed} at line {line}, column {column}.")

def _parse_img_attrs(attr_str: str, line: int, column: int) -> dict:
    attrs = _parse_tag_attrs(attr_str, line, column)
    if 'src' not in attrs:
        raise ValueError(f"<img> requires src=... at line {line}, column {column}.")
    if 'w' in attrs:
        attrs['w'] = _normalize_img_dimension(attrs['w'], 'w', line, column)
    if 'h' in attrs:
        attrs['h'] = _normalize_img_dimension(attrs['h'], 'h', line, column)
    if 'm' in attrs:
        attrs['m'] = _normalize_img_dimension(attrs['m'], 'm', line, column, allow_auto=False)
    return attrs

def _render_img_html(attrs: dict, forced_height: int = None, expand_cell: bool = False) -> str:
    src = _html.escape(str(attrs.get('src', '')), quote=True)
    html_attrs = [f'src="{src}"']

    width = attrs.get('w')
    height = attrs.get('h')
    if forced_height is not None:
        width = 'auto'
        height = str(forced_height)

    style_parts = ["display:inline-block", "vertical-align:middle"]
    if not expand_cell:
        style_parts.insert(0, "max-width:100%")

    if width and width != 'auto':
        html_attrs.append(f'width="{_html.escape(width, quote=True)}"')
        style_parts.append(f"width:{width}px")
    elif height and height != 'auto':
        style_parts.append("width:auto")

    if height and height != 'auto':
        html_attrs.append(f'height="{_html.escape(height, quote=True)}"')
        style_parts.append(f"height:{height}px")
    else:
        style_parts.append("height:auto")

    margin = attrs.get('m')
    if margin and margin != 'auto':
        style_parts.append(f"margin:{margin}px")

    for key, value in attrs.items():
        if key in ('src', 'w', 'h', 'm'):
            continue
        if _re.fullmatch(r'[a-zA-Z_:][-a-zA-Z0-9_:.]*', key):
            html_attrs.append(f'{key}="{_html.escape(str(value), quote=True)}"')

    html_attrs.append(f'style="{";".join(style_parts)};"')
    return f'<img {" ".join(html_attrs)}>'

def _scan_data_content(content: str):
    """
    Scans the data(...) inner string and extracts tokens.
    Returns a list of Token objects with tracking of source line and column.
    """
    tokens = []
    i = 0
    n = len(content)
    line = 1
    column = 1
    
    def advance(steps=1):
        nonlocal i, line, column
        for _ in range(steps):
            if i >= n:
                break
            if content[i] == '\n':
                line += 1
                column = 1
            else:
                column += 1
            i += 1

    while i < n:
        # Check for csv(...) calls
        csv_fn_match = _re.match(r'(?i)^csv\(', content[i:])
        if csv_fn_match:
            tok_line, tok_col = line, column
            start_i = i
            advance(len(csv_fn_match.group(0)))
            depth = 1
            in_q = False
            q_char = None
            while i < n and depth > 0:
                c = content[i]
                if not in_q:
                    if c in ('"', "'"):
                        in_q = True
                        q_char = c
                    elif c == '(':
                        depth += 1
                    elif c == ')':
                        depth -= 1
                else:
                    if c == q_char:
                        in_q = False
                advance(1)
            tokens.append(_Token(_TokenType.TEXT, content[start_i:i], tok_line, tok_col))
            continue

        # Check for generic tags: <tag_name>, <tag_name=attr_value>, or </tag_name>
        if content[i] == '<':
            # Closing tag
            tag_close_match = _re.match(r'^</([a-zA-Z_]+)>', content[i:])
            if tag_close_match:
                tok_line, tok_col = line, column
                tag_name = tag_close_match.group(1)
                advance(len(tag_close_match.group(0)))
                tokens.append(_Token(_TokenType.TAG_CLOSE, tag_name, tok_line, tok_col))
                continue
                
            # Opening tag
            tag_open = _scan_open_tag(content[i:])
            if tag_open:
                tok_line, tok_col = line, column
                tag_text, tag_name, tag_attr = tag_open
                advance(len(tag_text))
                tok_type = _TokenType.TAG_SELF if _is_self_tag(tag_name) else _TokenType.TAG_OPEN
                tokens.append(_Token(tok_type, tag_name, tok_line, tok_col, tag_attr))
                continue

            tokens.append(_Token(_TokenType.TEXT, '<', line, column))
            advance(1)
            continue

        # Check for string literals (double quotes)
        if content[i] == '"':
            tok_line, tok_col = line, column
            val_chars = []
            advance(1)  # skip open quote
            while i < n and content[i] != '"':
                val_chars.append(content[i])
                advance(1)
            if i < n:
                advance(1)  # skip close quote
            tokens.append(_Token(_TokenType.TEXT, "".join(val_chars), tok_line, tok_col))
            continue

        # Check for string literals (single quotes)
        if content[i] == "'":
            tok_line, tok_col = line, column
            val_chars = []
            advance(1)
            while i < n and content[i] != "'":
                val_chars.append(content[i])
                advance(1)
            if i < n:
                advance(1)
            tokens.append(_Token(_TokenType.TEXT, "".join(val_chars), tok_line, tok_col))
            continue

        # Semicolon
        if content[i] == ';':
            tokens.append(_Token(_TokenType.SEMICOLON, ';', line, column))
            advance(1)
            continue

        # Comma
        if content[i] == ',':
            tokens.append(_Token(_TokenType.COMMA, ',', line, column))
            advance(1)
            continue

        # Text/variable identifiers
        tok_line, tok_col = line, column
        val_chars = []
        while i < n and content[i] not in (',', ';', '<', '"', "'"):
            val_chars.append(content[i])
            advance(1)
        tokens.append(_Token(_TokenType.TEXT, "".join(val_chars), tok_line, tok_col))

    # Clean and filter empty text tokens
    cleaned_tokens = []
    for tok in tokens:
        if tok.type == _TokenType.TEXT:
            cleaned_val = tok.value.strip()
            if not cleaned_val and tok.value == "":
                cleaned_tokens.append(_Token(_TokenType.TEXT, "", tok.line, tok.column))
            elif cleaned_val:
                cleaned_tokens.append(_Token(_TokenType.TEXT, cleaned_val, tok.line, tok.column))
        else:
            cleaned_tokens.append(tok)
            
    return cleaned_tokens

def _scan_csv_cell_content(cell_val: str):
    """
    Scans a single CSV cell value, extracting tags (<tag>, <tag=val>, </tag>)
    and treating the rest of the text as a single text token.
    """
    tokens = []
    i = 0
    n = len(cell_val)
    line = 1
    column = 1
    current_text = []
    tok_line, tok_col = 1, 1

    def advance(steps=1):
        nonlocal i, line, column
        for _ in range(steps):
            if i >= n:
                break
            if cell_val[i] == '\n':
                line += 1
                column = 1
            else:
                column += 1
            i += 1

    def flush_text():
        if current_text:
            tokens.append(_Token(_TokenType.TEXT, "".join(current_text), tok_line, tok_col))
            current_text.clear()

    while i < n:
        if cell_val[i] == '<':
            # Closing tag
            tag_close_match = _re.match(r'^</([a-zA-Z_]+)>', cell_val[i:])
            if tag_close_match:
                flush_text()
                tag_line, tag_col = line, column
                tag_name = tag_close_match.group(1)
                advance(len(tag_close_match.group(0)))
                tokens.append(_Token(_TokenType.TAG_CLOSE, tag_name, tag_line, tag_col))
                tok_line, tok_col = line, column
                continue
                
            # Opening tag
            tag_open = _scan_open_tag(cell_val[i:])
            if tag_open:
                flush_text()
                tag_line, tag_col = line, column
                tag_text, tag_name, tag_attr = tag_open
                advance(len(tag_text))
                tok_type = _TokenType.TAG_SELF if _is_self_tag(tag_name) else _TokenType.TAG_OPEN
                tokens.append(_Token(tok_type, tag_name, tag_line, tag_col, tag_attr))
                tok_line, tok_col = line, column
                continue
        
        if not current_text:
            tok_line, tok_col = line, column
        current_text.append(cell_val[i])
        advance(1)
        
    flush_text()
    return tokens

def _load_csv(source_path: str):
    """
    Loads CSV data from a URL or local file path.
    Returns a list of lists of strings.
    """
    source_path = source_path.strip().strip('"\'')
    if source_path.startswith(('http://', 'https://')):
        try:
            req = _urllib_request.Request(
                source_path, 
                headers={'User-Agent': 'Mozilla/5.0 (tentags)'}
            )
            with _urllib_request.urlopen(req) as response:
                content = response.read().decode('utf-8')
            f = _io.StringIO(content)
            reader = _csv.reader(f)
            return list(reader)
        except Exception as e:
            return [[f"Error loading CSV URL: {e}"]]
    else:
        try:
            with open(source_path, mode='r', encoding='utf-8') as f:
                reader = _csv.reader(f)
                return list(reader)
        except Exception as e:
            return [[f"Error loading CSV file: {e}"]]

def _parse_data_arg(content: str, context: dict = None):
    """
    Parses data block content into a 2D grid of CellDesc.
    Supports variable interpolation using context dictionary.
    Expands csv("path_or_url") links dynamically into the token stream.
    Supports generic tags and style mapping.
    """
    if context is None:
        context = {}
        
    tokens = _scan_data_content(content)
    
    # Expand CSV tokens dynamically
    expanded_tokens = []
    for tok in tokens:
        if tok.type == _TokenType.TEXT:
            csv_match = _re.match(r'(?is)^csv\((.*)\)$', tok.value)
            if csv_match:
                csv_path = csv_match.group(1).strip().strip('"\'')
                csv_data = _load_csv(csv_path)
                
                csv_tokens = []
                for r_idx, row in enumerate(csv_data):
                    if r_idx > 0:
                        csv_tokens.append(_Token(_TokenType.SEMICOLON, ';', tok.line, tok.column))
                    for c_idx, cell_val in enumerate(row):
                        if c_idx > 0:
                            csv_tokens.append(_Token(_TokenType.COMMA, ',', tok.line, tok.column))
                        csv_tokens.extend(_scan_csv_cell_content(cell_val))
                expanded_tokens.extend(csv_tokens)
                continue
        expanded_tokens.append(tok)

    cells_grid = [[]]
    current_cell = CellDesc()

    active_tags = []
    cm_counter = 0
    rm_counter = 0

    def apply_active_tags(cell):
        for active_tok in active_tags:
            tag_lower = active_tok.value.lower()
            if tag_lower == 'cm':
                cell.cm_block_id = getattr(active_tok, 'block_id', None)
            elif tag_lower == 'rm':
                cell.rm_block_id = getattr(active_tok, 'block_id', None)
            elif tag_lower == 'b':
                cell.styles['font-weight'] = 'bold'
            elif tag_lower == 'i':
                cell.styles['font-style'] = 'italic'
            elif tag_lower == 'u':
                decorations = cell.styles.get('text-decoration', '').split()
                if 'underline' not in decorations:
                    decorations.append('underline')
                cell.styles['text-decoration'] = ' '.join(decorations).strip()
            elif tag_lower == 's':
                decorations = cell.styles.get('text-decoration', '').split()
                if 'line-through' not in decorations:
                    decorations.append('line-through')
                cell.styles['text-decoration'] = ' '.join(decorations).strip()
            elif tag_lower == 'color':
                if active_tok.attr:
                    cell.styles['color'] = active_tok.attr
            elif tag_lower == 'bg':
                if active_tok.attr:
                    cell.styles['background-color'] = active_tok.attr
            elif tag_lower == 'url':
                if active_tok.attr:
                    link = _parse_link(active_tok.attr)
                    cell.link = link
                    if link.scheme != "goto":
                        cell.styles['href'] = active_tok.attr
            elif tag_lower == 'fs':
                if active_tok.attr:
                    cell.styles['font-size'] = active_tok.attr if any(c.isalpha() or c == '%' for c in active_tok.attr) else f"{active_tok.attr}px"
            elif tag_lower == 'left':
                cell.styles['text-align'] = 'left'
            elif tag_lower == 'center':
                cell.styles['text-align'] = 'center'
            elif tag_lower == 'right':
                cell.styles['text-align'] = 'right'

    # Apply initial active tags if any exist (though usually stack is empty at start)
    apply_active_tags(current_cell)

    def commit_cell():
        nonlocal current_cell
        raw_text = "".join(current_cell.text_parts).strip()
        
        # Apply variable interpolation context
        if raw_text in context:
            raw_text = str(context[raw_text])
            
        current_cell.raw_expr = raw_text
        
        cells_grid[-1].append(current_cell)
        current_cell = CellDesc()
        # Carry over active tags spanning cell boundaries
        apply_active_tags(current_cell)

    for tok in expanded_tokens:
        if tok.type == _TokenType.TAG_SELF:
            tag_name = tok.value.lower()
            if tag_name == 'img':
                current_cell.images.append(_parse_img_attrs(tok.attr, tok.line, tok.column))
                apply_active_tags(current_cell)
            elif tag_name == 'mark':
                mark = str(tok.attr or "").strip()
                if not mark:
                    raise ValueError(f"<mark> requires a name at line {tok.line}, column {tok.column}.")
                current_cell.mark = mark
                apply_active_tags(current_cell)
            elif tag_name == 'value':
                if not tok.attr:
                    raise ValueError(f"<value> requires an address at line {tok.line}, column {tok.column}.")
                address = _parse_address(tok.attr)
                placeholder = f"__TENTAGS_VALUE_REF_{len(current_cell.value_refs)}__"
                current_cell.value_refs.append((placeholder, address))
                current_cell.text_parts.append(placeholder)
                apply_active_tags(current_cell)
            else:
                current_cell.text_parts.append(f"<{tok.value}>")

        elif tok.type == _TokenType.TAG_OPEN:
            tag_name = tok.value.lower()
            if tag_name == 'cm' and any(t.value.lower() == 'cm' for t in active_tags):
                raise ValueError(f"Nested <cm> blocks are not supported at line {tok.line}, column {tok.column}.")
            if tag_name == 'rm' and any(t.value.lower() == 'rm' for t in active_tags):
                raise ValueError(f"Nested <rm> blocks are not supported at line {tok.line}, column {tok.column}.")
            
            # Assign unique block IDs for merge tags
            if tag_name == 'cm':
                cm_counter += 1
                tok.block_id = cm_counter
            elif tag_name == 'rm':
                rm_counter += 1
                tok.block_id = rm_counter
                
            active_tags.append(tok)
            apply_active_tags(current_cell)
            
        elif tok.type == _TokenType.TAG_CLOSE:
            tag_name = tok.value.lower()
            # Find matching tag in stack
            match_idx = -1
            for idx in reversed(range(len(active_tags))):
                if active_tags[idx].value.lower() == tag_name:
                    match_idx = idx
                    break
            if match_idx == -1:
                raise ValueError(f"Unexpected closing tag </{tag_name}> at line {tok.line}, column {tok.column}.")
            active_tags.pop(match_idx)
            # Re-apply remaining active styles
            apply_active_tags(current_cell)
            
        elif tok.type == _TokenType.COMMA:
            commit_cell()
        elif tok.type == _TokenType.SEMICOLON:
            commit_cell()
            cells_grid.append([])
        elif tok.type == _TokenType.TEXT:
            apply_active_tags(current_cell)
            current_cell.text_parts.append(tok.value)

    # Commit final cell
    commit_cell()

    # Check for unclosed tags
    if active_tags:
        unclosed = active_tags[-1]
        raise ValueError(f"Missing closing tag </{unclosed.value}> (opened at line {unclosed.line}, column {unclosed.column}).")

    def is_plain_empty_cell(cell):
        return (
            cell.raw_expr == ""
            and not cell.styles
            and not cell.images
            and cell.link is None
            and cell.mark is None
            and not cell.value_refs
            and cell.cm_block_id is None
            and cell.rm_block_id is None
            and cell.border_flags == BorderFlags.NONE
        )

    # Remove trailing empty row if created by a trailing semicolon.
    # Keep styled-empty cells: style(...) often uses empty cell bodies to carry presentation.
    if len(cells_grid) > 1:
        last_row = cells_grid[-1]
        if len(last_row) == 1 and is_plain_empty_cell(last_row[0]):
            cells_grid.pop()

    # Apply Column Merge borders
    for row in cells_grid:
        i = 0
        n = len(row)
        while i < n:
            if row[i].cm_block_id is None:
                i += 1
                continue

            start_idx = i
            while i < n and row[i].cm_block_id is not None:
                i += 1
            end_idx = i - 1

            if start_idx < end_idx:
                row[start_idx].border_flags |= BorderFlags.HIDE_RIGHT
                for mid in range(start_idx + 1, end_idx):
                    row[mid].border_flags |= (BorderFlags.HIDE_LEFT | BorderFlags.HIDE_RIGHT)
                row[end_idx].border_flags |= BorderFlags.HIDE_LEFT

    # Apply Row Merge borders
    max_cols = max(len(row) for row in cells_grid) if cells_grid else 0
    num_rows = len(cells_grid)
    for c in range(max_cols):
        r = 0
        while r < num_rows:
            if c >= len(cells_grid[r]):
                r += 1
                continue

            if cells_grid[r][c].rm_block_id is None:
                r += 1
                continue

            start_row = r
            while r < num_rows and c < len(cells_grid[r]) and cells_grid[r][c].rm_block_id is not None:
                r += 1
            end_row = r - 1

            if start_row < end_row:
                cells_grid[start_row][c].border_flags |= BorderFlags.HIDE_BOTTOM
                for mid in range(start_row + 1, end_row):
                    cells_grid[mid][c].border_flags |= (BorderFlags.HIDE_TOP | BorderFlags.HIDE_BOTTOM)
                cells_grid[end_row][c].border_flags |= BorderFlags.HIDE_TOP

    return cells_grid

def _parse_args_string(arg_str: str):
    """
    Parses the full arguments string into parameters and the raw data block.
    """
    parts = []
    current = []
    depth = 0
    in_quote = False
    quote_char = None
    i = 0
    n = len(arg_str)

    while i < n:
        c = arg_str[i]
        if not in_quote:
            if c in ('"', "'"):
                in_quote = True
                quote_char = c
                current.append(c)
            elif c == '(':
                depth += 1
                current.append(c)
            elif c == ')':
                depth -= 1
                current.append(c)
            elif c == ',' and depth == 0:
                parts.append("".join(current).strip())
                current = []
            else:
                current.append(c)
        else:
            if c == quote_char:
                in_quote = False
            current.append(c)
        i += 1
        
    if current:
        parts.append("".join(current).strip())

    return parts

def _extract_preamble_scale(
    args: list[str],
    rows: int,
    cols: int,
) -> tuple[list[str], dict[int, int], dict[int, int]]:
    scale_indexes = [
        index
        for index, value in enumerate(args)
        if str(value).strip().lower().startswith("scale(")
    ]
    if not scale_indexes:
        return args, {}, {}
    if len(scale_indexes) > 1:
        raise ScaleError("Preamble may contain only one scale(...) extension.")

    scale_index = scale_indexes[0]
    if scale_index < 6:
        raise ScaleError("scale(...) must follow the six required preamble arguments.")

    first_content_index = next(
        (
            index
            for index, value in enumerate(args)
            if str(value).strip().lower().startswith(("style(", "data("))
        ),
        None,
    )
    if first_content_index is not None and scale_index > first_content_index:
        raise ScaleError("scale(...) belongs to the preamble and must appear before style(...) and data(...).")

    row_scales, col_scales = _parse_scale_block(args[scale_index], rows, cols)
    remaining = args[:scale_index] + args[scale_index + 1:]
    return remaining, row_scales, col_scales

def _validate_vertical_scale_base(row_scales: dict[int, int], cell_height: int) -> None:
    if row_scales and cell_height <= 0:
        raise ScaleError("Vertical scale requires cell_height greater than 0.")

def _overlay_style_and_data(style_grid: list[list[CellDesc]], data_grid: list[list[CellDesc]]) -> list[list[CellDesc]]:
    """
    Overlays a presentation style grid of cells onto a content data grid of cells.
    Returns a new merged 2D grid of CellDesc.
    """
    cells_grid = []
    max_r = max(len(style_grid), len(data_grid))
    for r in range(max_r):
        row_cells = []
        style_row_len = len(style_grid[r]) if r < len(style_grid) else 0
        data_row_len = len(data_grid[r]) if r < len(data_grid) else 0
        max_c = max(style_row_len, data_row_len)
        
        for c in range(max_c):
            merged_cell = CellDesc()
            style_cell = style_grid[r][c] if r < len(style_grid) and c < len(style_grid[r]) else None
            data_cell = data_grid[r][c] if r < len(data_grid) and c < len(data_grid[r]) else None

            if style_cell:
                merged_cell.cm_block_id = style_cell.cm_block_id
                merged_cell.rm_block_id = style_cell.rm_block_id
                merged_cell.border_flags = style_cell.border_flags
                merged_cell.styles = dict(style_cell.styles)
                merged_cell.images = list(style_cell.images)
                merged_cell.link = style_cell.link
                merged_cell.mark = style_cell.mark
                merged_cell.value_refs = list(style_cell.value_refs)

            if data_cell:
                merged_cell.raw_expr = data_cell.raw_expr
                merged_cell.text_parts = list(data_cell.text_parts)
                merged_cell.images.extend(data_cell.images)
                merged_cell.value_refs.extend(data_cell.value_refs)
                if data_cell.link is not None:
                    merged_cell.link = data_cell.link
                if data_cell.mark is not None:
                    merged_cell.mark = data_cell.mark
                if data_cell.cm_block_id is not None:
                    merged_cell.cm_block_id = data_cell.cm_block_id
                if data_cell.rm_block_id is not None:
                    merged_cell.rm_block_id = data_cell.rm_block_id
                merged_cell.border_flags |= data_cell.border_flags
                for k, v in data_cell.styles.items():
                    merged_cell.styles[k] = v

            row_cells.append(merged_cell)
        cells_grid.append(row_cells)
    return cells_grid

def parse(formula_args: str, context: dict = None) -> TableModel:
    """
    Parses formula arguments string, style and data blocks into a TableModel.
    Supports both legacy format (only data) and new format (style and data).
    """
    args = _parse_args_string(formula_args)
    if len(args) < 6:
        raise ValueError("TABLE requires at least 6 arguments (rows, cols, border_width, border_color, border_style, stretch)")

    rows = int(args[0])
    cols = int(args[1])
    border_width = int(args[2])
    border_color = args[3].strip('"\'')
    border_style = args[4].strip('"\'')
    stretch = int(args[5])
    args, row_scales, col_scales = _extract_preamble_scale(args, rows, cols)

    cell_height = 30
    style_block_str = None
    data_block_str = None

    rem = args[6:]
    if len(rem) == 1:
        data_block_str = rem[0]
    elif len(rem) == 2:
        if rem[0].lower().startswith('style('):
            style_block_str = rem[0]
            data_block_str = rem[1]
        else:
            cell_height = int(rem[0])
            data_block_str = rem[1]
    elif len(rem) >= 3:
        cell_height = int(rem[0])
        style_block_str = rem[1]
        data_block_str = rem[2]

    _validate_vertical_scale_base(row_scales, cell_height)

    cells_grid = []

    if style_block_str and data_block_str:
        style_inner = ""
        match_style = _re.match(r'(?is)^style\((.*)\)$', style_block_str)
        if match_style:
            style_inner = match_style.group(1).strip()
        style_grid = _parse_data_arg(style_inner, context)

        data_inner = ""
        match_data = _re.match(r'(?is)^data\((.*)\)$', data_block_str)
        if match_data:
            data_inner = match_data.group(1).strip()
        data_grid = _parse_data_arg(data_inner, context)

        cells_grid = _overlay_style_and_data(style_grid, data_grid)
    elif data_block_str:
        match = _re.match(r'(?is)^data\((.*)\)$', data_block_str)
        if match:
            data_inner = match.group(1).strip()
            cells_grid = _parse_data_arg(data_inner, context)

    model = TableModel(
        rows=rows,
        cols=cols,
        cells=cells_grid,
        border_width=border_width,
        border_color=border_color,
        border_style=border_style,
        stretch=stretch,
        cell_height=cell_height,
        row_scales=row_scales,
        col_scales=col_scales,
    )
    return _resolve_value_refs(model)

def render_html(
    model: TableModel,
    address_resolver: AddressResolver = None,
    address_context: AddressTarget = None,
) -> str:
    """
    Renders a TableModel into an HTML table string.
    """
    build_mark_index(model)
    current_target = address_context or _make_address_target(model)
    resolver = address_resolver or _local_address_resolver(current_target)

    border_style = str(model.border_style).strip().lower()
    apply_to_outer = True
    apply_to_inner = False
    if border_style.endswith("-1"):
        apply_to_inner = True
        apply_to_outer = True
        border_style = border_style[:-2]
    elif border_style.endswith("-0"):
        apply_to_inner = False
        apply_to_outer = False
        border_style = border_style[:-2]

    table_border = f"border:{model.border_width}px {border_style} {model.border_color};" if apply_to_outer else ""
    table_style = f"border-collapse:collapse;{table_border}"
    has_images = any(
        bool(cell.images)
        for row in model.cells
        for cell in row
    )
    if model.stretch == 1:
        if has_images:
            table_style += "width:auto;table-layout:auto;"
        else:
            table_style += "width:100%;height:100%;table-layout:fixed;"
    else:
        table_style += "width:100%;table-layout:fixed;"

    html = [f'<table style="{table_style}">']

    if model.col_scales and model.cols > 0:
        col_weights = [model.col_scales.get(col, 1) for col in range(model.cols)]
        total_weight = sum(col_weights)
        html.append("<colgroup>")
        for weight in col_weights:
            percentage = 100.0 * weight / total_weight
            width_text = f"{percentage:.6f}".rstrip("0").rstrip(".")
            html.append(f'<col style="width:{width_text}%">')
        html.append("</colgroup>")

    td_border = f"border:{model.border_width}px {border_style} {model.border_color};" if apply_to_inner else ""

    for r in range(model.rows):
        row_scale = model.row_scales.get(r, 1)
        scaled_cell_height = model.cell_height * row_scale
        if model.stretch == 0:
            row_height_style = f"height:{scaled_cell_height}px;"
        elif not has_images:
            if model.row_scales:
                row_weight_total = sum(model.row_scales.get(row, 1) for row in range(model.rows))
                percentage = 100.0 * row_scale / row_weight_total if row_weight_total else 0
                height_text = f"{percentage:.6f}".rstrip("0").rstrip(".")
                row_height_style = f"height:{height_text}%;"
            else:
                row_height_style = f"height:{100.0 / model.rows}%;" if model.rows else ""
        elif row_scale > 1:
            row_height_style = f"height:{scaled_cell_height}px;"
        else:
            row_height_style = ""
        html.append(f'<tr style="{row_height_style}">')
        for c in range(model.cols):
            val = ""
            href = None
            link = None
            mark = None
            images = []
            cell_styles = {}
            border_overrides = []

            if r < len(model.cells) and c < len(model.cells[r]):
                cell = model.cells[r][c]
                val = cell.raw_expr
                if val == 'None':
                    val = ''
                images = cell.images
                link = cell.link
                mark = cell.mark
                cell_styles = cell.styles
                
                if cell.border_flags & BorderFlags.HIDE_LEFT:
                    border_overrides.append("border-left:none;")
                if cell.border_flags & BorderFlags.HIDE_RIGHT:
                    border_overrides.append("border-right:none;")
                if cell.border_flags & BorderFlags.HIDE_TOP:
                    border_overrides.append("border-top:none;")
                if cell.border_flags & BorderFlags.HIDE_BOTTOM:
                    border_overrides.append("border-bottom:none;")

                # Serialize inline styles (excluding href which is rendered as anchor tag)
                for prop, prop_val in cell.styles.items():
                    if prop == 'href':
                        href = prop_val
                    else:
                        border_overrides.append(f"{prop}:{prop_val};")

                if link is not None:
                    href = _link_to_html_href(link, resolver, current_target)

            if "text-align" not in cell_styles:
                border_overrides.append("text-align:center;")
            border_overrides.append("vertical-align:middle;")

            overrides_css = "".join(border_overrides)
            td_style = f"{td_border}padding:0;{overrides_css}"
            
            if model.stretch == 0:
                td_style += f"height:{scaled_cell_height}px;"

            forced_img_height = scaled_cell_height if model.stretch == 0 and images else None
            content = val + "".join(
                _render_img_html(img, forced_height=forced_img_height, expand_cell=model.stretch == 1)
                for img in images
            )

            if mark:
                content = f'<span id="{_html_mark_id(mark, current_target.html_prefix)}"></span>{content}'

            # Wrap with anchor if url tag was used
            if href:
                safe_href = _html.escape(str(href), quote=True)
                content = f'<a href="{safe_href}" style="color:inherit;text-decoration:inherit;">{content}</a>'

            html.append(f'<td id="{_html_cell_id(r, c, current_target.html_prefix)}" style="{td_style}">{content}</td>')
        html.append('</tr>')

    html.append('</table>')
    return "".join(html)

def _normalize_color_to_hex(color_str: str) -> str:
    color_str = color_str.strip().strip('"\'').lower()
    
    color_map = {
        'black': '000000', 'white': 'ffffff', 'red': 'ff0000',
        'green': '008000', 'blue': '0000ff', 'yellow': 'ffff00',
        'gray': '808080', 'grey': '808080', 'silver': 'c0c0c0',
        'maroon': '800000', 'purple': '800080', 'fuchsia': 'ff00ff',
        'lime': '00ff00', 'olive': '808000', 'navy': '000080',
        'teal': '008080', 'aqua': '00ffff', 'orange': 'ffa500'
    }
    
    if color_str in color_map:
        return color_map[color_str]
        
    hex_val = color_str.replace('#', '')
    if len(hex_val) == 3:
        return "".join(c * 2 for c in hex_val)
    elif len(hex_val) in (6, 8):
        return hex_val
        
    return '000000'

def _write_model_to_sheet(
    model: TableModel,
    ws,
    start_row: int = 1,
    address_resolver: AddressResolver = None,
    address_context: AddressTarget = None,
):
    from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
    from openpyxl.utils.units import DEFAULT_COLUMN_WIDTH
    mark_index = build_mark_index(model)
    current_target = address_context or _make_address_target(
        model,
        xlsx_sheet_name=ws.title,
        xlsx_start_row=start_row,
    )
    resolver = address_resolver or _local_address_resolver(current_target)
    
    # Configure borders mapping
    border_style = str(model.border_style).strip().lower()
    apply_to_outer = True
    apply_to_inner = False
    if border_style.endswith("-1"):
        apply_to_inner = True
        apply_to_outer = True
        border_style = border_style[:-2]
    elif border_style.endswith("-0"):
        apply_to_inner = False
        apply_to_outer = False
        border_style = border_style[:-2]

    excel_border_style = 'thin'
    if border_style == 'dashed':
        excel_border_style = 'dashed'
    elif border_style == 'dotted':
        excel_border_style = 'dotted'
    elif model.border_width > 1:
        excel_border_style = 'medium'
        
    excel_color = _normalize_color_to_hex(model.border_color)
    border_side = Side(style=excel_border_style, color=excel_color)

    if model.col_scales:
        default_column_width = ws.sheet_format.defaultColWidth or DEFAULT_COLUMN_WIDTH
        for col, scale in model.col_scales.items():
            column_name = _column_to_name(col)
            desired_width = default_column_width * scale
            current_width = ws.column_dimensions[column_name].width
            ws.column_dimensions[column_name].width = max(current_width or 0, desired_width)
    
    # Write values, borders, alignments, and heights
    for r in range(model.rows):
        row_scale = model.row_scales.get(r, 1)
        if model.stretch == 0 or row_scale > 1:
            ws.row_dimensions[start_row + r].height = model.cell_height * row_scale
            
        for c in range(model.cols):
            cell_ref = ws.cell(row=start_row + r, column=c + 1)
            
            val = ""
            cell = None
            cell_styles = {}
            images = []
            link = None
            if r < len(model.cells) and c < len(model.cells[r]):
                cell = model.cells[r][c]
                val = cell.raw_expr
                if val == 'None':
                    val = ''
                cell_styles = cell.styles
                images = cell.images
                link = cell.link
            cell_ref.value = val or (images[0].get('src', '') if images else '')

            if images:
                img_src = images[0].get('src', '')
                if img_src.startswith(('http://', 'https://')):
                    cell_ref.hyperlink = img_src
                elif _os.path.exists(img_src):
                    try:
                        from openpyxl.drawing.image import Image as XLImage

                        sheet_img = XLImage(img_src)
                        width = images[0].get('w')
                        height = images[0].get('h')
                        original_width = sheet_img.width
                        original_height = sheet_img.height

                        if width and width != 'auto' and height and height != 'auto':
                            sheet_img.width = int(width)
                            sheet_img.height = int(height)
                        elif width and width != 'auto':
                            sheet_img.width = int(width)
                            if original_width:
                                sheet_img.height = int(original_height * int(width) / original_width)
                        elif height and height != 'auto':
                            sheet_img.height = int(height)
                            if original_height:
                                sheet_img.width = int(original_width * int(height) / original_height)

                        ws.add_image(sheet_img, cell_ref.coordinate)
                        if not val:
                            cell_ref.value = None
                    except Exception:
                        pass
            
            if apply_to_inner:
                left_s = border_side
                right_s = border_side
                top_s = border_side
                bottom_s = border_side
            elif apply_to_outer:
                left_s = border_side if c == 0 else Side()
                right_s = border_side if c == model.cols - 1 else Side()
                top_s = border_side if r == 0 else Side()
                bottom_s = border_side if r == model.rows - 1 else Side()
            else:
                left_s = Side()
                right_s = Side()
                top_s = Side()
                bottom_s = Side()

            if cell is not None:
                if cell.border_flags & BorderFlags.HIDE_LEFT:
                    left_s = Side()
                if cell.border_flags & BorderFlags.HIDE_RIGHT:
                    right_s = Side()
                if cell.border_flags & BorderFlags.HIDE_TOP:
                    top_s = Side()
                if cell.border_flags & BorderFlags.HIDE_BOTTOM:
                    bottom_s = Side()

            cell_ref.border = Border(
                left=left_s,
                right=right_s,
                top=top_s,
                bottom=bottom_s,
            )

            h_align = cell_styles.get('text-align', 'center')
            cell_ref.alignment = Alignment(horizontal=h_align, vertical='center', wrap_text=True)

            # Apply font styles (bold, italic, color, size)
            is_bold = cell_styles.get('font-weight') == 'bold'
            is_italic = cell_styles.get('font-style') == 'italic'
            font_color = None
            if 'color' in cell_styles:
                font_color = _normalize_color_to_hex(cell_styles['color'])
            font_size = None
            if 'font-size' in cell_styles:
                num_match = _re.search(r'\d+', cell_styles['font-size'])
                if num_match:
                    font_size = int(num_match.group(0))
                
            decorations = cell_styles.get('text-decoration', '').split()
            is_underline = 'underline' in decorations
            is_strike = 'line-through' in decorations
                
            if is_bold or is_italic or font_color or font_size or is_underline or is_strike:
                cell_ref.font = Font(
                    bold=is_bold, 
                    italic=is_italic, 
                    color=font_color,
                    size=font_size,
                    underline="single" if is_underline else None,
                    strike=is_strike
                )
                
            # Apply background color
            if 'background-color' in cell_styles:
                bg_color = _normalize_color_to_hex(cell_styles['background-color'])
                cell_ref.fill = PatternFill(
                    start_color=bg_color, 
                    end_color=bg_color, 
                    fill_type='solid'
                )

            # Apply hyperlink if url tag was used
            xlsx_href = (
                _link_to_xlsx_hyperlink(link, ws.title, mark_index, resolver, current_target)
                if link is not None
                else cell_styles.get('href')
            )
            if xlsx_href:
                cell_ref.hyperlink = xlsx_href
                # Apply standard hyperlink styling if no color already set
                if 'color' not in cell_styles:
                    cell_ref.font = Font(
                        bold=is_bold, italic=is_italic,
                        size=font_size,
                        underline='single',
                        color='0563C1'  # Excel default hyperlink blue
                    )

def render_xlsx(
    model: TableModel,
    filepath_or_stream,
    address_resolver: AddressResolver = None,
    address_context: AddressTarget = None,
):
    """
    Renders a TableModel into an Excel (.xlsx) file.
    Requires openpyxl to be installed.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("The 'openpyxl' package is required for Excel rendering. Please install it using 'pip install openpyxl'.")

    wb = openpyxl.Workbook()
    # Remove default active sheet
    wb.remove(wb.active)
    ws = wb.create_sheet(title=address_context.xlsx_sheet_name if address_context and address_context.xlsx_sheet_name else "Table")
    _write_model_to_sheet(
        model,
        ws,
        start_row=1,
        address_resolver=address_resolver,
        address_context=address_context,
    )
    wb.save(filepath_or_stream)

def _create_pdf_table_object(
    model: TableModel,
    address_resolver: AddressResolver = None,
    address_context: AddressTarget = None,
    available_width: float = None,
):
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib.styles import ParagraphStyle
    current_target = address_context or _make_address_target(model)
    resolver = address_resolver or _local_address_resolver(current_target)
    pdf_fonts = _pdf_font_names()

    # Use the same named-color and HEX normalization as the XLSX renderer.
    def to_rl_color(color_value: str, default=colors.black):
        if not color_value:
            return default
        try:
            normalized = _normalize_color_to_hex(str(color_value))
            return colors.HexColor('#' + normalized[-6:])
        except Exception:
            return default

    data_matrix = []
    table_styles = []

    # Configure grid borders
    border_style = str(model.border_style).strip().lower()
    apply_to_outer = True
    apply_to_inner = False
    if border_style.endswith("-1"):
        apply_to_inner = True
        apply_to_outer = True
        border_style = border_style[:-2]
    elif border_style.endswith("-0"):
        apply_to_inner = False
        apply_to_outer = False
        border_style = border_style[:-2]

    border_w = max(0.5, float(model.border_width))
    border_c = to_rl_color(model.border_color)
    border_dash = None
    if border_style == 'dashed':
        border_dash = [4, 3]
    elif border_style == 'dotted':
        border_dash = [1, 2]

    def pdf_line(command, start, end):
        return (command, start, end, border_w, border_c, 1, border_dash)

    if apply_to_inner:
        table_styles.append(pdf_line('BOX', (0, 0), (-1, -1)))

        for r in range(model.rows):
            for c in range(1, model.cols):
                left_cell = model.cells[r][c - 1] if r < len(model.cells) and c - 1 < len(model.cells[r]) else None
                right_cell = model.cells[r][c] if r < len(model.cells) and c < len(model.cells[r]) else None
                hidden = (
                    (left_cell is not None and left_cell.border_flags & BorderFlags.HIDE_RIGHT)
                    or (right_cell is not None and right_cell.border_flags & BorderFlags.HIDE_LEFT)
                )
                if not hidden:
                    table_styles.append(pdf_line('LINEBEFORE', (c, r), (c, r)))

        for r in range(1, model.rows):
            for c in range(model.cols):
                top_cell = model.cells[r - 1][c] if r - 1 < len(model.cells) and c < len(model.cells[r - 1]) else None
                bottom_cell = model.cells[r][c] if r < len(model.cells) and c < len(model.cells[r]) else None
                hidden = (
                    (top_cell is not None and top_cell.border_flags & BorderFlags.HIDE_BOTTOM)
                    or (bottom_cell is not None and bottom_cell.border_flags & BorderFlags.HIDE_TOP)
                )
                if not hidden:
                    table_styles.append(pdf_line('LINEABOVE', (c, r), (c, r)))
    elif apply_to_outer:
        table_styles.append(pdf_line('BOX', (0, 0), (-1, -1)))
        
    table_styles.append(('VALIGN', (0, 0), (-1, -1), 'MIDDLE'))

    for r in range(model.rows):
        row_data = []
        for c in range(model.cols):
            val = ""
            cell_styles = {}
            images = []
            link = None
            mark = None
            if r < len(model.cells) and c < len(model.cells[r]):
                cell = model.cells[r][c]
                val = cell.raw_expr
                if val == 'None':
                    val = ''
                cell_styles = cell.styles
                images = cell.images
                link = cell.link
                mark = cell.mark
            if images and not val:
                val = images[0].get('src', '')

            # Background fill
            if 'background-color' in cell_styles:
                bg_c = to_rl_color(cell_styles['background-color'])
                table_styles.append(('BACKGROUND', (c, r), (c, r), bg_c))

            # Alignment
            h_align = cell_styles.get('text-align', 'center').upper()
            if h_align in ('LEFT', 'CENTER', 'RIGHT'):
                table_styles.append(('ALIGN', (c, r), (c, r), h_align))

            # Typography formatting
            is_bold = cell_styles.get('font-weight') == 'bold'
            is_italic = cell_styles.get('font-style') == 'italic'
            font_color = to_rl_color(cell_styles.get('color', ''), default=colors.black)
            decorations_pdf = cell_styles.get('text-decoration', '').split()
            is_underline = 'underline' in decorations_pdf
            is_strike = 'line-through' in decorations_pdf
            
            font_size = 11
            if 'font-size' in cell_styles:
                num_match = _re.search(r'\d+', cell_styles['font-size'])
                if num_match:
                    font_size = int(num_match.group(0))

            font_name = pdf_fonts["regular"]
            if is_bold and is_italic:
                font_name = pdf_fonts["bold_italic"]
            elif is_bold:
                font_name = pdf_fonts["bold"]
            elif is_italic:
                font_name = pdf_fonts["italic"]

            table_styles.append(('FONTNAME', (c, r), (c, r), font_name))
            table_styles.append(('FONTSIZE', (c, r), (c, r), font_size))
            table_styles.append(('TEXTCOLOR', (c, r), (c, r), font_color))
            if is_underline:
                table_styles.append(('UNDERLINE', (c, r), (c, r)))
            if is_strike:
                table_styles.append(('STRIKETHROUGH', (c, r), (c, r)))

            href_pdf = (
                _link_to_html_href(link, resolver, current_target, "pdf_prefix")
                if link is not None
                else cell_styles.get('href')
            )
            if href_pdf and 'color' not in cell_styles:
                # Default hyperlink blue
                table_styles.append(('TEXTCOLOR', (c, r), (c, r), colors.HexColor('#0563C1')))
                table_styles.append(('UNDERLINE', (c, r), (c, r)))

            if val != '':
                rl_align = 1 # center
                if h_align == 'LEFT':
                    rl_align = 0
                elif h_align == 'RIGHT':
                    rl_align = 2
                
                p_style = ParagraphStyle(
                    f'Cell_{r}_{c}',
                    fontName=font_name,
                    fontSize=font_size,
                    leading=font_size + 3,
                    textColor=font_color,
                    alignment=rl_align
                )
                anchors = [f'<a name="{_html_cell_id(r, c, current_target.pdf_prefix)}"/>']
                if mark:
                    anchors.append(f'<a name="{_html_mark_id(mark, current_target.pdf_prefix)}"/>')
                cell_text = "".join(anchors) + str(val)
                if is_underline:
                    cell_text = f'<u>{cell_text}</u>'
                if is_strike:
                    cell_text = f'<strike>{cell_text}</strike>'
                if href_pdf:
                    cell_text = f'<link href="{href_pdf}">{cell_text}</link>'
                row_data.append(Paragraph(cell_text, p_style))
            else:
                row_data.append("")
        data_matrix.append(row_data)

    col_widths = None
    if model.col_scales and available_width and model.cols > 0:
        col_weights = [model.col_scales.get(col, 1) for col in range(model.cols)]
        total_weight = sum(col_weights)
        col_widths = [available_width * weight / total_weight for weight in col_weights]

    row_heights = None
    min_row_heights = None
    if model.stretch == 0 and model.cell_height > 0:
        row_heights = [
            max(25, model.cell_height * model.row_scales.get(row, 1))
            for row in range(model.rows)
        ]
    elif model.row_scales and model.cell_height > 0:
        min_row_heights = [
            model.cell_height * model.row_scales.get(row, 1)
            if model.row_scales.get(row, 1) > 1
            else 0
            for row in range(model.rows)
        ]

    return Table(
        data_matrix,
        colWidths=col_widths,
        rowHeights=row_heights,
        minRowHeights=min_row_heights,
        style=TableStyle(table_styles),
    )

def render_pdf(
    model: TableModel,
    filepath_or_stream: _Union[str, _Any],
    address_resolver: AddressResolver = None,
    address_context: AddressTarget = None,
) -> None:
    """
    Renders a TableModel directly to a PDF document using ReportLab.
    Translates IR coordinates, merged regions, background fills, fonts, and borders into native ReportLab TableStyles.
    """
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate
    except ImportError:
        raise ImportError(
            "The 'reportlab' package is required for PDF rendering. "
            "Install it via 'pip install tentags[pdf]' or 'pip install reportlab'."
        )

    doc = SimpleDocTemplate(
        filepath_or_stream,
        pagesize=landscape(letter) if model.cols > 4 else letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    table = _create_pdf_table_object(
        model,
        address_resolver=address_resolver,
        address_context=address_context,
        available_width=doc.width,
    )
    doc.build([table])

def _load_style(filepath_or_str: _Union[str, list], context: dict = None) -> list[list[CellDesc]]:
    if not isinstance(filepath_or_str, str):
        return filepath_or_str
    
    content = filepath_or_str
    if _os.path.exists(filepath_or_str):
        with open(filepath_or_str, 'r', encoding='utf-8') as f:
            content = f.read()
    
    content = content.strip()
    match = _re.match(r'(?is)^style\((.*)\)$', content)
    if match:
        content = match.group(1).strip()
    return _parse_data_arg(content, context)

def _load_data(filepath_or_str: _Union[str, list], context: dict = None) -> list[list[CellDesc]]:
    if not isinstance(filepath_or_str, str):
        return filepath_or_str
    
    content = filepath_or_str
    if _os.path.exists(filepath_or_str):
        with open(filepath_or_str, 'r', encoding='utf-8') as f:
            content = f.read()
    
    content = content.strip()
    match = _re.match(r'(?is)^data\((.*)\)$', content)
    if match:
        content = match.group(1).strip()
    return _parse_data_arg(content, context)

def _csv_load(filepath_or_url: str) -> list[list[CellDesc]]:
    raw_data = _load_csv(filepath_or_url)
    grid = []
    for r in raw_data:
        row = []
        for val in r:
            cell = CellDesc()
            cell.raw_expr = val.strip()
            cell.text_parts = [cell.raw_expr]
            row.append(cell)
        grid.append(row)
    return grid

def compile(preamble: _Any, style: _Any, data: _Any, context: dict = None) -> TableModel:
    """
    Compiles TenTags style and data sources into an intermediate representation (TableModel).
    
    The returned TableModel can later be rendered into HTML, PDF, or XLSX formats.
    """
    style_grid = _load_style(style, context)
    data_grid = _load_data(data, context)
    
    cells_grid = _overlay_style_and_data(style_grid, data_grid)
    
    rows = len(cells_grid)
    cols = max(len(row) for row in cells_grid) if cells_grid else 0
    border_width = 1
    border_color = "#cbd5e1"
    border_style = "solid"
    stretch = 0
    cell_height = 30
    row_scales = {}
    col_scales = {}

    if preamble:
        if isinstance(preamble, str):
            p_args = _parse_args_string(preamble)
            if len(p_args) >= 6:
                rows = int(p_args[0])
                cols = int(p_args[1])
                p_args, row_scales, col_scales = _extract_preamble_scale(p_args, rows, cols)
                border_width = int(p_args[2])
                border_color = p_args[3].strip('"\'')
                border_style = p_args[4].strip('"\'')
                stretch = int(p_args[5])
                if len(p_args) >= 7:
                    cell_height = int(p_args[6])
                _validate_vertical_scale_base(row_scales, cell_height)
            elif len(p_args) >= 3:
                border_width = int(p_args[0])
                border_color = p_args[1].strip('"\'')
                border_style = p_args[2].strip('"\'')
                if len(p_args) >= 4:
                    stretch = int(p_args[3])
                if len(p_args) >= 5:
                    cell_height = int(p_args[4])
        elif isinstance(preamble, dict):
            border_width = preamble.get('border_width', border_width)
            border_color = preamble.get('border_color', border_color)
            border_style = preamble.get('border_style', border_style)
            stretch = preamble.get('stretch', stretch)
            cell_height = preamble.get('cell_height', cell_height)
            rows = preamble.get('rows', rows)
            cols = preamble.get('cols', cols)
            scale_block = _serialize_scale_mapping(preamble.get('scale'), rows, cols)
            if scale_block:
                row_scales, col_scales = _parse_scale_block(scale_block, rows, cols)
            _validate_vertical_scale_base(row_scales, cell_height)

    model = TableModel(
        rows=rows,
        cols=cols,
        cells=cells_grid,
        border_width=border_width,
        border_color=border_color,
        border_style=border_style,
        stretch=stretch,
        cell_height=cell_height,
        row_scales=row_scales,
        col_scales=col_scales,
    )
    return _resolve_value_refs(model)

def render(preamble_or_formula: _Any, style: _Any = None, data: _Any = None, context: dict = None) -> str:
    """
    Renders a TenTags table directly to an HTML string.

    Supports two calling styles:
    
    1. Render a self-contained layout formula:
       import tentags
       html = tentags.render('3,3,1,"black","solid",0, data(A,B,C; D,E,F; G,H,I)')

    2. Render with decoupled style templates, data blocks, and preambles:
       import tentags
       preamble = '3, 3, 1, "blue", "solid-1"'
       style = 'style(<bg=white><center><b><cm>Title, , </cm></b></center></bg>; A, B, C)'
       data = 'data( , , ; Value A, Value B, Value C)'
       html = tentags.render(preamble, style, data)

    Parameters:
        preamble_or_formula: Either a self-contained TenTags formula string, or a layout preamble.
        style: Optional style block (string or list of cells).
        data: Optional data block (string or list of cells).
        context: Optional dictionary of variable substitutions.

    Returns:
        A formatted <table>...</table> HTML string.
    """
    if isinstance(style, dict) and data is None and context is None:
        context = style
        style = None

    try:
        if style is not None and data is not None:
            model = compile(preamble_or_formula, style, data, context)
        else:
            model = parse(preamble_or_formula, context)
        return render_html(model)
    except Exception as e:
        return str(e)

def multitable_html(
    tables: list, 
    layout: str = _SENTINEL, 
    cols: int = _SENTINEL, 
    gap: str = _SENTINEL, 
    full_page: bool = _SENTINEL, 
    context: dict = None,
    settings: dict = None,
) -> str:
    """
    Assembles and renders multiple tables into a single HTML string.
    """
    settings = _html_settings(
        DEFAULT_MULTITABLE_HTML_SETTINGS,
        settings,
        {
            "layout": layout,
            "cols": cols,
            "gap": gap,
            "full_page": full_page,
        },
    )
    tables = _prepare_multitable_items(tables, settings)
    resolver = AddressResolver()
    materialized = []
    for index, item in enumerate(tables):
        model = _table_item_model(item, context)
        meta = _table_item_meta(item, index)
        address_context = AddressContext(
            project=meta["project"],
            document=meta["document"],
            list_name=meta["list"],
        )
        prefix = AddressResolver.html_prefix_for(address_context, index)
        target = resolver.register(
            model,
            project=meta["project"],
            document=meta["document"],
            list_name=meta["list"],
            html_prefix=prefix,
            xlsx_sheet_name=meta["sheet_name"],
            pdf_prefix=prefix,
        )
        materialized.append((model, meta, target))

    _validate_multitable_columns([(model, meta) for model, meta, _ in materialized], settings)

    rendered_tables = []
    for model, meta, target in materialized:
        table_html = render_html(model, address_resolver=resolver, address_context=target)
        title = meta["title"]
        if title:
            table_html = f"<div><h3>{title}</h3>{table_html}</div>"
        rendered_tables.append(table_html)

    if settings["layout"] == "grid":
        style_container = f"display: grid; grid-template-columns: repeat({settings['cols']}, 1fr); gap: {settings['gap']};"
    else:
        style_container = f"display: flex; flex-direction: column; gap: {settings['gap']};"

    container = f'<div style="{style_container}">\n' + "\n".join(rendered_tables) + '\n</div>'

    if settings["full_page"]:
        html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>{_html.escape(str(settings["html_title"]))}</title>
</head>
<body>
    {container}
</body>
</html>"""
    else:
        html = container
    _write_html_output(settings.get("output"), html)
    return html

def multitable_xlsx(
    tables: list, 
    filepath_or_stream=None, 
    mode: str = _SENTINEL, 
    gap: int = _SENTINEL, 
    show_titles: bool = _SENTINEL, 
    context: dict = None,
    settings: dict = None,
) -> None:
    """
    Assembles and renders multiple tables into a single Excel (.xlsx) workbook.
    """
    settings = _xlsx_settings(
        DEFAULT_MULTITABLE_XLSX_SETTINGS,
        settings,
        {
            "mode": mode,
            "gap": gap,
            "show_titles": show_titles,
        },
    )
    filepath_or_stream = _normalize_output_target(filepath_or_stream or settings.get("output"))
    if filepath_or_stream is None:
        raise ValueError("multitable_xlsx requires filepath_or_stream or settings['output'].")
    tables = _prepare_multitable_items(tables, settings)

    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        raise ImportError("The 'openpyxl' package is required for Excel rendering. Please install it using 'pip install openpyxl'.")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    materialized = [
        (_table_item_model(item, context), _table_item_meta(item, index))
        for index, item in enumerate(tables)
    ]
    _validate_multitable_columns(materialized, settings)

    if settings["mode"] == "sheets":
        resolver = AddressResolver()
        targets = []
        for model, meta in materialized:
            target = resolver.register(
                model,
                project=meta["project"],
                document=meta["document"],
                list_name=meta["list"],
                xlsx_sheet_name=meta["sheet_name"],
                xlsx_start_row=1,
            )
            targets.append(target)

        for (model, meta), target in zip(materialized, targets):
            sheet_name = meta["sheet_name"]
            ws = wb.create_sheet(title=sheet_name)
            _write_model_to_sheet(
                model,
                ws,
                start_row=1,
                address_resolver=resolver,
                address_context=target,
            )
    else: # stacked
        stacked_sheet_name = settings["stacked_sheet_name"]
        ws = wb.create_sheet(title=stacked_sheet_name)
        current_row = 1
        resolver = AddressResolver()
        targets = []
        for model, meta in materialized:
            title = meta["title"]
            table_start_row = current_row + 1 if settings["show_titles"] and title else current_row
            target = resolver.register(
                model,
                project=meta["project"],
                document=meta["document"],
                list_name=meta["list"],
                xlsx_sheet_name=stacked_sheet_name,
                xlsx_start_row=table_start_row,
            )
            targets.append(target)
            current_row = table_start_row + model.rows + settings["gap"]

        current_row = 1
        for (model, meta), target in zip(materialized, targets):
            title = meta["title"]
            if settings["show_titles"] and title:
                cell = ws.cell(row=current_row, column=1)
                cell.value = title
                cell.font = Font(bold=True, size=14)
                current_row += 1
            
            _write_model_to_sheet(
                model,
                ws,
                start_row=current_row,
                address_resolver=resolver,
                address_context=target,
            )
            current_row += model.rows + settings["gap"]

    wb.save(filepath_or_stream)

def multitable_pdf(
    tables: list, 
    filepath_or_stream=None, 
    page_size: str = _SENTINEL, 
    orientation: str = _SENTINEL, 
    page_break_after_each: bool = _SENTINEL, 
    margins: tuple = _SENTINEL, 
    tables_per_row: _Any = _SENTINEL,
    tables_per_page: _Any = _SENTINEL,
    gap: int = _SENTINEL,
    context: dict = None,
    settings: dict = None,
) -> None:
    """
    Assembles and renders multiple tables into a single PDF document.
    """
    settings = _pdf_settings(
        DEFAULT_MULTITABLE_PDF_SETTINGS,
        settings,
        {
            "page_size": page_size,
            "orientation": orientation,
            "page_break_after_each": page_break_after_each,
            "margins": margins,
            "tables_per_row": tables_per_row,
            "tables_per_page": tables_per_page,
            "gap": gap,
        },
    )
    filepath_or_stream = _normalize_output_target(filepath_or_stream or settings.get("output"))
    if filepath_or_stream is None:
        raise ValueError("multitable_pdf requires filepath_or_stream or settings['output'].")
    tables = _prepare_multitable_items(tables, settings)

    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, PageBreak, Paragraph, Table, TableStyle
        from reportlab.lib.styles import ParagraphStyle
    except ImportError:
        raise ImportError(
            "The 'reportlab' package is required for PDF rendering. "
            "Install it via 'pip install tentags[pdf]' or 'pip install reportlab'."
        )

    # Resolve page size
    base_page_size = letter
    if str(settings["page_size"]).lower() == "a4":
        from reportlab.lib.pagesizes import A4
        base_page_size = A4

    actual_page_size = landscape(base_page_size) if str(settings["orientation"]).lower() == "landscape" else base_page_size

    doc = SimpleDocTemplate(
        filepath_or_stream,
        pagesize=actual_page_size,
        leftMargin=settings["margins"][0],
        rightMargin=settings["margins"][1],
        topMargin=settings["margins"][2],
        bottomMargin=settings["margins"][3]
    )
    pdf_fonts = _pdf_font_names()

    resolver = AddressResolver()
    materialized = []
    for index, item in enumerate(tables):
        model = _table_item_model(item, context)
        meta = _table_item_meta(item, index)
        address_context = AddressContext(
            project=meta["project"],
            document=meta["document"],
            list_name=meta["list"],
        )
        prefix = AddressResolver.html_prefix_for(address_context, index)
        target = resolver.register(
            model,
            project=meta["project"],
            document=meta["document"],
            list_name=meta["list"],
            xlsx_sheet_name=meta["sheet_name"],
            html_prefix=prefix,
            pdf_prefix=prefix,
        )
        materialized.append((model, meta, target))

    _validate_multitable_columns([(model, meta) for model, meta, _ in materialized], settings)

    def table_block(model, meta, target, index, available_width=None):
        block = []
        title = meta["title"]
        if title:
            p_style = ParagraphStyle(
                f"Title_{index}",
                fontName=pdf_fonts["bold"],
                fontSize=14,
                leading=18,
                spaceAfter=10
            )
            block.append(Paragraph(title, p_style))
        block.append(
            _create_pdf_table_object(
                model,
                address_resolver=resolver,
                address_context=target,
                available_width=available_width,
            )
        )
        return block

    def auto_tables_per_row(blocks, content_width, gap_size):
        if not blocks:
            return 1
        widths = []
        for block in blocks:
            table_flowable = block[-1] if block else None
            if table_flowable is None:
                continue
            width, _ = table_flowable.wrap(content_width, actual_page_size[1])
            widths.append(width)
        if not widths:
            return 1
        max_width = max(widths)
        if max_width <= 0:
            return 1
        columns = int((content_width + gap_size) // (max_width + gap_size))
        return max(1, min(columns, len(blocks)))

    def block_height(block, width):
        total = 0
        for flowable in block:
            _, height = flowable.wrap(width, actual_page_size[1])
            total += height
        return total

    def auto_tables_per_page(blocks, tables_per_row, col_width, content_height, gap_size):
        if not blocks:
            return 1
        row_heights = []
        for row_index in range(0, len(blocks), tables_per_row):
            row_blocks = blocks[row_index:row_index + tables_per_row]
            row_heights.append(max(block_height(block, col_width) for block in row_blocks))
        if not row_heights:
            return 1

        rows_that_fit = 0
        used_height = 0
        for row_height in row_heights:
            next_height = used_height + row_height
            if rows_that_fit > 0:
                next_height += gap_size
            if rows_that_fit > 0 and next_height > content_height:
                break
            used_height = next_height
            rows_that_fit += 1
            if used_height >= content_height:
                break

        return max(1, rows_that_fit) * tables_per_row

    story = []
    preliminary_blocks = [
        table_block(model, meta, target, index)
        for index, (model, meta, target) in enumerate(materialized)
    ]
    content_width = actual_page_size[0] - settings["margins"][0] - settings["margins"][1]
    content_height = actual_page_size[1] - settings["margins"][2] - settings["margins"][3]
    gap_size = settings.get("gap") or 0
    tables_per_row = settings.get("tables_per_row") or 1
    if tables_per_row == "auto":
        tables_per_row = auto_tables_per_row(preliminary_blocks, content_width, gap_size)
    tables_per_page = settings.get("tables_per_page")
    col_width = (content_width - (tables_per_row - 1) * gap_size) / tables_per_row
    table_width = col_width if tables_per_row == 1 else max(1, col_width - gap_size)
    blocks = [
        table_block(model, meta, target, index, available_width=table_width)
        for index, (model, meta, target) in enumerate(materialized)
    ]
    if tables_per_page == "auto":
        tables_per_page = auto_tables_per_page(
            blocks,
            tables_per_row,
            col_width,
            content_height,
            gap_size,
        )
    if tables_per_page is not None:
        tables_per_row = min(tables_per_row, tables_per_page)

    if tables_per_row > 1:
        page_size = tables_per_page or len(blocks)
        page_groups = [
            blocks[index:index + page_size]
            for index in range(0, len(blocks), page_size)
        ]

        for page_index, page_blocks in enumerate(page_groups):
            grid_rows = []
            for row_index in range(0, len(page_blocks), tables_per_row):
                row_blocks = page_blocks[row_index:row_index + tables_per_row]
                while len(row_blocks) < tables_per_row:
                    row_blocks.append("")
                grid_rows.append(row_blocks)

            grid_table = Table(
                grid_rows,
                colWidths=[col_width] * tables_per_row,
                style=TableStyle([
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), gap_size / 2),
                    ("RIGHTPADDING", (0, 0), (-1, -1), gap_size / 2),
                    ("TOPPADDING", (0, 0), (-1, -1), gap_size / 2),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), gap_size / 2),
                ]),
            )
            story.append(grid_table)
            if page_index < len(page_groups) - 1:
                story.append(PageBreak())
    else:
        for i, block in enumerate(blocks):
            story.extend(block)

            should_break = False
            if tables_per_page is not None:
                should_break = (i + 1) % tables_per_page == 0
            else:
                should_break = bool(settings["page_break_after_each"])
            if should_break and i < len(tables) - 1:
                story.append(PageBreak())

    doc.build(story)

def features() -> dict:
    """
    Returns the availability of optional render backends depending on local dependencies.

    Example output:
    {
        "html": True,
        "pdf": True,
        "xlsx": False
    }
    """
    has_pdf = False
    try:
        from reportlab.platypus import SimpleDocTemplate
        has_pdf = True
    except ImportError:
        pass

    has_xlsx = False
    try:
        import openpyxl
        has_xlsx = True
    except ImportError:
        pass

    return {
        "html": True,
        "pdf": has_pdf,
        "xlsx": has_xlsx
    }

def info() -> None:
    """
    Prints package version, available renderers, optional dependencies,
    and system diagnostic information to the console.
    """
    import sys
    feats = features()
    
    print(f"TenTags {__version__}\n")
    print(f"Author       {__author__}")
    print(f"License      {__license__}")
    print(f"Website      {__homepage__}\n")
    
    print(f"Python       {sys.version.split()[0]}\n")
    
    print("Features")
    print("────────")
    print(f"{'✓' if feats['html'] else '✗'} HTML")
    print(f"{'✓' if feats['pdf'] else '✗'} PDF")
    print(f"{'✓' if feats['xlsx'] else '✗'} XLSX")
    print("✓ Validation")
    print("✓ Demo")

def get_prompt(print_output: bool = False) -> str:
    """
    Return the bundled TenTags LLM bootstrap prompt.

    Pass print_output=True to also print the prompt.
    """
    from importlib import resources as _resources
    resource_name = "TENTAGS_LLM_BOOTSTRAP_PROMPT.md"

    try:
        try:
            prompt = (
                _resources.files(__package__)
                .joinpath(resource_name)
                .read_text(encoding="utf-8")
            )
        except AttributeError:
            with _resources.open_text(__package__, resource_name, encoding="utf-8") as f:
                prompt = f.read()
    except (FileNotFoundError, ModuleNotFoundError):
        prompt_path = _os.path.join(_os.path.dirname(_os.path.dirname(__file__)), resource_name)
        with open(prompt_path, "r", encoding="utf-8") as f:
            prompt = f.read()

    if print_output:
        print(prompt)
    return prompt

def validate(formula: str) -> dict:
    """
    Validates a TenTags formula's syntax and tag balance.
    Returns a dict with 'status' ('ok' or 'error') and a descriptive message.
    """
    if not isinstance(formula, str):
        return {"status": "error", "message": "Formula must be a string."}
        
    try:
        parse(formula)
        return {"status": "ok", "message": "Syntax OK"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

def demo(name: str = "dashboard", filepath_prefix: str = "demo") -> None:
    """
    Generates a demonstration report to showcase TenTags output formats.
    Supported names: 'dashboard', 'invoice', 'table'.
    """
    name = name.lower()
    if name == "dashboard":
        preamble = '4, 4, 1, "#cbd5e1", "solid", 0, 45'
        style = 'style(<fs=18><bg=#1e293b><color=white><b><cm>Q3 Financial Performance Dashboard, , , , </cm></b></color></bg></fs>; <bg=#f1f5f9><b><left>Department</left></b></bg>, <bg=#f1f5f9><b><center>Revenue</center></b></bg>, <bg=#f1f5f9><b><center>Expenses</center></b></bg>, <bg=#f1f5f9><b><center>Net Profit</center></b></bg>; <left>Engineering</left>, <right>"$240,000"</right>, <right>"$180,000"</right>, <bg=#dcfce7><color=#166534><b><right>"+$60,000"</right></b></color></bg>; <left>Sales & Marketing</left>, <right>"$310,000"</right>, <right>"$210,000"</right>, <bg=#dcfce7><color=#166534><b><right>"+$100,000"</right></b></color></bg>)'
        data = 'data()'
    elif name == "invoice":
        preamble = '7, 4, 1, "#e2e8f0", "solid-1", 1, 30'
        style = 'style(<fs=16><b><left><cm>INVOICE #1024, , , </cm></left></b></fs>; <left><cm>Date: 2026-07-14, , , </cm></left>; <bg=#3b82f6><color=white><b>Item</b></color></bg>, <bg=#3b82f6><color=white><b>Quantity</b></color></bg>, <bg=#3b82f6><color=white><b>Price</b></color></bg>, <bg=#3b82f6><color=white><b>Total</b></color></bg>; ; ; ; <right><b><cm>Grand Total: , , </cm></b></right>, <b>$1,650</b>)'
        data = 'data( , , , ; , , , ; Premium Wood Logs, 10, $150, $1,500; Steel Beams, 5, $30, $150; , , , )'
    else:
        preamble = '3, 3, 1, "green", "solid-1", 1'
        style = 'style(<bg=white><center><b><cm>Test Table, , </cm></b></center></bg>; <center><b>Header 1</b></center>, <center><b>Header 2</b></center>, <center><b>Header 3</b></center>; <left>A1</left>, <center>B1</center>, <right>C1</right>)'
        data = 'data()'

    model = compile(style=style, data=data, preamble=preamble)
    
    html_content = render_html(model)
    with open(f"{filepath_prefix}_{name}.html", "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f"Generated HTML demo: {filepath_prefix}_{name}.html")

    feats = features()
    if feats["xlsx"]:
        render_xlsx(model, f"{filepath_prefix}_{name}.xlsx")
        print(f"Generated Excel demo: {filepath_prefix}_{name}.xlsx")
        
    if feats["pdf"]:
        render_pdf(model, f"{filepath_prefix}_{name}.pdf")
        print(f"Generated PDF demo: {filepath_prefix}_{name}.pdf")
